from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

_DANGEROUS_SPREADSHEET_PREFIXES = ("=", "+", "-", "@")


def decode_csv_bytes(data: bytes) -> str | None:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return None


def parse_csv_cells(data: bytes) -> list[str]:
    decoded = decode_csv_bytes(data)
    if decoded is None:
        return []

    values: list[str] = []
    reader = csv.reader(io.StringIO(decoded))
    for row in reader:
        for cell in row:
            candidate = cell.strip()
            if candidate:
                values.append(candidate)
    return values


def parse_xlsx_cells(data: bytes) -> list[str]:
    """Extract all non-empty cell values from an xlsx file."""
    try:
        workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        values: list[str] = []
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        candidate = str(cell).strip()
                        if candidate:
                            values.append(candidate)
        workbook.close()
        return values
    except Exception:
        return []


def parse_spreadsheet_cells(data: bytes, filename: str) -> list[str]:
    """Parse cells from either CSV or XLSX file based on filename."""
    if filename.lower().endswith(".xlsx"):
        return parse_xlsx_cells(data)
    else:
        return parse_csv_cells(data)


def sanitize_csv_cell(value: object) -> str:
    text = str(value or "")
    if text.startswith(_DANGEROUS_SPREADSHEET_PREFIXES):
        return f"'{text}"
    return text


def build_csv_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([sanitize_csv_cell(header) for header in headers])
    for row in rows:
        writer.writerow([sanitize_csv_cell(cell) for cell in row])
    return buffer.getvalue().encode("utf-8")
